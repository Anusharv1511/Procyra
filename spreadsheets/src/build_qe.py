"""Build QE_Toolkit.xlsx — Quality Engineering companion workbook."""
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.worksheet.datavalidation import DataValidation
from common import (build_guide, build_setup, build_workflow, header_row, put_input,
                    put_formula, put_link, put_label, note, title, SETUP)

wb = Workbook()
wb.remove(wb.active)

build_guide(wb, "QE Toolkit (Quality Engineering)", [
    ("Setup", "Industry & process type — drives the defect terminology and thresholds everywhere."),
    ("Workflow - Start Here", "Guided DMAIC sequence, mirroring the web app's playbook."),
    ("NC Log", "Non-conformance log — one row per occurrence, consistent codes."),
    ("Pareto", "Auto-sorted Pareto from the NC Log: RANK-based ordering, cumulative % computed on the sorted series."),
    ("FMEA Register", "S x O x D = RPN with action flags vs your industry threshold, plus an auto-sorted top-risk table."),
    ("CAPA Tracker", "Corrective/preventive actions with due dates and overdue flags."),
], honesty_notes=[
    "The Pareto's category list is typed by you (column A) because auto-extracting unique codes needs functions (UNIQUE) that break older Excel/LibreOffice. The web app aggregates codes automatically.",
    "The web app auto-drafts a CAPA when a defect code repeats past your industry threshold; in Excel, watch the repeat counts on the Pareto and open CAPAs yourself.",
    "Charts reference the full input range; unfilled rows can render as gaps or zeros — a charting artifact, not data.",
])
build_setup(wb)
build_workflow(wb, [
    ("Define", "Agree one defect coding list for the team; write the problem statement.", "NC Log", "Free-typed codes (SCRATCH vs Scratch vs scratched) split one problem into three bars."),
    ("Measure", "Log every occurrence as it happens with code, area, and quantity.", "NC Log", "End-of-shift memory logging undercounts small defects."),
    ("Analyze", "Read the Pareto: attack the biggest bar, check the cumulative 80% point.", "Pareto", "A flat Pareto (no dominant bar) suggests a systemic cause, not a defect-specific one."),
    ("Improve", "Open a CAPA for the chosen defect; assign owner and due date.", "CAPA Tracker", "A CAPA without an owner and a date is a wish."),
    ("Control", "Update FMEA occurrence scores for addressed failure modes; keep logging.", "FMEA Register", "If RPN doesn't drop after the fix, the fix didn't work — or the scoring is optimistic."),
])

# ---------------------------------------------------------------- NC Log
ws = wb.create_sheet("NC Log")
title(ws, "A1", "Non-conformance log")
put_label(ws, "A3", "Term for this industry:")
put_link(ws, "B3", f"={SETUP['defect']}")
HDR = 5; FIRST = 6; LAST = 45  # 40 rows
header_row(ws, HDR, 1, ["Date", "Code", "Process area", "Qty", "Severity", "Description"],
           widths=[11, 14, 16, 6, 10, 40])
EX = [
    ("2026-06-01", "SCRATCH", "Line 4", 3, "Minor", "Surface scratches after conveyor transfer"),
    ("2026-06-01", "LEAK-01", "Test bench", 1, "Major", "Seal leak at 2 bar"),
    ("2026-06-02", "SCRATCH", "Line 4", 2, "Minor", ""),
    ("2026-06-03", "MISALIGN", "Line 2", 1, "Major", "Bracket hole misaligned"),
    ("2026-06-04", "SCRATCH", "Line 4", 4, "Minor", "After fixture change"),
    ("2026-06-05", "DENT", "Receiving", 2, "Minor", "Packaging damage inbound"),
    ("2026-06-08", "LEAK-01", "Test bench", 1, "Major", ""),
    ("2026-06-09", "SCRATCH", "Line 4", 2, "Minor", ""),
    ("2026-06-10", "WRONG-LABEL", "Packout", 5, "Critical", "Label mixup, contained"),
    ("2026-06-11", "MISALIGN", "Line 2", 2, "Major", ""),
    ("2026-06-12", "SCRATCH", "Line 4", 3, "Minor", ""),
    ("2026-06-15", "DENT", "Receiving", 1, "Minor", ""),
    ("2026-06-16", "LEAK-01", "Test bench", 2, "Major", "Recurring seal supplier lot"),
    ("2026-06-17", "SCRATCH", "Line 4", 2, "Minor", ""),
    ("2026-06-18", "MISALIGN", "Line 2", 1, "Major", ""),
    ("2026-06-19", "DENT", "Receiving", 1, "Minor", ""),
]
dv = DataValidation(type="list", formula1='"Minor,Major,Critical"', allow_blank=True)
ws.add_data_validation(dv)
for i in range(40):
    r = FIRST + i
    row = EX[i] if i < len(EX) else ("", "", "", "", "", "")
    for j, col in enumerate("ABCDEF"):
        put_input(ws, f"{col}{r}", row[j])
    dv.add(ws[f"E{r}"])
note(ws, "A47", "Keep codes consistent — the Pareto matches on exact text. Add rows INSIDE the range if you need more than 40.")

# ---------------------------------------------------------------- Pareto
ws = wb.create_sheet("Pareto")
title(ws, "A1", "Pareto — auto-sorted from the NC Log")
note(ws, "A2", "Type each distinct code once in column A. Counts pull from the NC Log; the sorted table and chart update automatically.")
HDR = 4; FIRST = 5; LAST = 16  # 12 category slots
header_row(ws, HDR, 1, ["Code (type once)", "Total qty", "Sort key", "Rank"],
           widths=[16, 10, 10, 7])
cats = ["SCRATCH", "LEAK-01", "MISALIGN", "DENT", "WRONG-LABEL"]
for i in range(12):
    r = FIRST + i
    put_input(ws, f"A{r}", cats[i] if i < len(cats) else "")
    put_link(ws, f"B{r}", f"=IF($A{r}=\"\",\"\",SUMIFS('NC Log'!$D$6:$D$45,'NC Log'!$B$6:$B$45,$A{r}))")
    # tiny row-based tiebreak keeps RANK unique when counts tie
    put_formula(ws, f"C{r}", f'=IF(ISNUMBER($B{r}),$B{r}+(100-ROW())/100000,"")')
    put_formula(ws, f"D{r}", f'=IF(ISNUMBER($C{r}),RANK($C{r},$C$5:$C$16),"")')
note(ws, "C3", "Sort key/Rank are helpers for the sorted table — leave them alone.")

put_label(ws, "F3", "Sorted (largest first) — chart reads this", bold=True, size=9)
header_row(ws, HDR, 6, ["#", "Code", "Qty", "Cumulative %"], widths=[4, 16, 8, 12])
for i in range(12):
    r = FIRST + i
    put_label(ws, f"F{r}", i + 1)
    put_formula(ws, f"G{r}", f'=IF(COUNTIF($D$5:$D$16,$F{r})=0,"",INDEX($A$5:$A$16,MATCH($F{r},$D$5:$D$16,0)))')
    put_formula(ws, f"H{r}", f'=IF(COUNTIF($D$5:$D$16,$F{r})=0,"",INDEX($B$5:$B$16,MATCH($F{r},$D$5:$D$16,0)))')
    c = put_formula(ws, f"I{r}", f'=IF(ISNUMBER($H{r}),SUM($H$5:$H{r})/SUM($H$5:$H$16),"")')
    c.number_format = "0.0%"

bar = BarChart(); bar.type = "col"; bar.title = "Pareto"; bar.height = 9; bar.width = 22
data = Reference(ws, min_col=8, min_row=HDR, max_row=LAST)
bar.add_data(data, titles_from_data=True)
bar.set_categories(Reference(ws, min_col=7, min_row=FIRST, max_row=LAST))
line = LineChart()
ldata = Reference(ws, min_col=9, min_row=HDR, max_row=LAST)
line.add_data(ldata, titles_from_data=True)
for s in line.series:
    s.smooth = False
line.y_axis.axId = 200
line.y_axis.title = "Cumulative %"
line.y_axis.crosses = "max"
bar.y_axis.title = "Qty"
bar += line
ws.add_chart(bar, "A19")
note(ws, "A18", "Bars are sorted descending; the cumulative % line is computed on the SORTED order so it rises monotonically. Empty category slots may render as gaps — charting artifact.")

# ---------------------------------------------------------------- FMEA Register
ws = wb.create_sheet("FMEA Register")
title(ws, "A1", "FMEA register — RPN with action flags")
put_label(ws, "A3", "RPN action threshold (from Setup):")
put_link(ws, "C3", f"={SETUP['rpn']}")
HDR = 5; FIRST = 6; LAST = 25  # 20 rows
header_row(ws, HDR, 1, ["Process step", "Failure mode", "Effect", "Cause", "S", "O", "D",
                        "RPN", "Action?", "Sort key", "Rank", "Recommended action", "Status"],
           widths=[16, 20, 20, 20, 4, 4, 4, 7, 9, 9, 6, 26, 10])
FEX = [
    ("Welding", "Cold weld", "Joint fails in service", "Low current setting", 8, 5, 4, "Poka-yoke current interlock", "Open"),
    ("Assembly", "Missing clip", "Rattle, warranty claim", "Manual step skipped", 6, 4, 3, "Add sensor check at station", "Open"),
    ("Packout", "Wrong label", "Ship to wrong customer", "Similar SKUs adjacent", 9, 3, 2, "Separate label stock bins", "Done"),
    ("Test", "Leak undetected", "Field failure", "Test pressure too low", 9, 2, 6, "Raise test pressure per spec", "Open"),
    ("Receiving", "Damaged housing accepted", "Scrap downstream", "No inbound check", 5, 4, 4, "Sample inspection on arrival", "None"),
]
dvs = DataValidation(type="list", formula1='"None,Open,Done"', allow_blank=True)
ws.add_data_validation(dvs)
for i in range(20):
    r = FIRST + i
    row = FEX[i] if i < len(FEX) else ("",) * 9
    for j, col in enumerate(["A", "B", "C", "D", "E", "F", "G"]):
        put_input(ws, f"{col}{r}", row[j])
    put_formula(ws, f"H{r}", f'=IF(COUNT($E{r}:$G{r})<3,"",$E{r}*$F{r}*$G{r})')
    put_formula(ws, f"I{r}", f'=IF(ISNUMBER($H{r}),IF($H{r}>=$C$3,"ACTION",""),"")')
    put_formula(ws, f"J{r}", f'=IF(ISNUMBER($H{r}),$H{r}+(100-ROW())/100000,"")')
    put_formula(ws, f"K{r}", f'=IF(ISNUMBER($J{r}),RANK($J{r},$J$6:$J$25),"")')
    put_input(ws, f"L{r}", row[7] if i < len(FEX) else "")
    put_input(ws, f"M{r}", row[8] if i < len(FEX) else "")
    dvs.add(ws[f"M{r}"])

put_label(ws, "O3", "Top 5 risks (auto-sorted)", bold=True, size=9)
header_row(ws, 5, 15, ["#", "Failure mode", "RPN"], widths=[4, 22, 7])
for i in range(5):
    r = 6 + i
    put_label(ws, f"O{r}", i + 1)
    put_formula(ws, f"P{r}", f'=IF(COUNTIF($K$6:$K$25,$O{r})=0,"",INDEX($B$6:$B$25,MATCH($O{r},$K$6:$K$25,0)))')
    put_formula(ws, f"Q{r}", f'=IF(COUNTIF($K$6:$K$25,$O{r})=0,"",INDEX($H$6:$H$25,MATCH($O{r},$K$6:$K$25,0)))')
note(ws, "A27", "RPN = S x O x D (1-10 each). ACTION appears when RPN meets your industry threshold. In the web app the register re-sorts itself and threshold crossings raise alerts automatically.")

# ---------------------------------------------------------------- CAPA Tracker
ws = wb.create_sheet("CAPA Tracker")
title(ws, "A1", "CAPA tracker")
HDR = 4; FIRST = 5; LAST = 24
header_row(ws, HDR, 1, ["ID", "Title", "Source", "Linked code", "Owner", "Opened", "Due",
                        "Status", "Days to due", "Overdue?"],
           widths=[6, 34, 9, 12, 12, 11, 11, 12, 10, 10])
CEX = [
    ("C-001", "Recurring SCRATCH on Line 4 — fixture redesign", "Auto", "SCRATCH", "J. Rivera", "2026-06-12", "2026-07-15", "In progress"),
    ("C-002", "Seal supplier lot control for LEAK-01", "Manual", "LEAK-01", "M. Chen", "2026-06-16", "2026-07-01", "Open"),
    ("C-003", "Label bin separation at packout", "Manual", "WRONG-LABEL", "A. Osei", "2026-06-10", "2026-06-20", "Closed"),
]
dvsrc = DataValidation(type="list", formula1='"Manual,Auto"', allow_blank=True)
dvst = DataValidation(type="list", formula1='"Draft,Open,In progress,Closed,Verified"', allow_blank=True)
ws.add_data_validation(dvsrc); ws.add_data_validation(dvst)
for i in range(20):
    r = FIRST + i
    row = CEX[i] if i < len(CEX) else ("",) * 8
    for j, col in enumerate(["A", "B", "C", "D", "E", "F", "G", "H"]):
        put_input(ws, f"{col}{r}", row[j])
    dvsrc.add(ws[f"C{r}"]); dvst.add(ws[f"H{r}"])
    put_formula(ws, f"I{r}", f'=IF(OR($G{r}="",$H{r}="Closed",$H{r}="Verified"),"",DATEVALUE($G{r})-TODAY())')
    put_formula(ws, f"J{r}", f'=IF(ISNUMBER($I{r}),IF($I{r}<0,"OVERDUE",""),"")')
note(ws, "A26", "Dates as YYYY-MM-DD text; Days-to-due uses DATEVALUE and TODAY(), so it updates every time you open the file. Source=Auto mirrors the web app's auto-drafted CAPAs.")

wb.save("../QE_Toolkit.xlsx")
print("QE_Toolkit.xlsx written")
