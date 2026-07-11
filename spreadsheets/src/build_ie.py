"""Build IE_Toolkit.xlsx — Industrial Engineering companion workbook."""
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series
from common import (build_guide, build_setup, build_workflow, header_row, put_input,
                    put_formula, put_link, put_label, note, title)

wb = Workbook()
wb.remove(wb.active)

build_guide(wb, "IE Toolkit (Industrial Engineering)", [
    ("Setup", "Industry & process type — adapts terminology used across sheets."),
    ("Workflow - Start Here", "Guided sequence for establishing a defensible standard time."),
    ("Time Study", "Element-level observed times, performance rating, PFD allowances -> normal and standard time."),
    ("Learning Curve", "Wright-model projection of per-unit time from the standard time and a learning rate."),
], honesty_notes=[
    "Performance rating is a judgment call by the observer — the sheet computes with it, it can't make it for you.",
    "The learning-curve projection models new operators / new products; it is not a steady-state standard.",
    "Charts reference the full projection range; unfilled rows can render as gaps or zeros — a charting artifact, not data.",
])
build_setup(wb)
build_workflow(wb, [
    ("Define", "Pick the job and break it into 5-12 repeatable elements with clear start/stop points.", "Time Study", "Elements longer than ~30s hide variation; shorter than ~2s are hard to time."),
    ("Measure", "Record 5-10 observed cycles per element; rate the operator's pace (100 = normal).", "Time Study", "Rate the pace you SAW, not the pace you want."),
    ("Analyze", "Review normal times; set Personal/Fatigue/Delay allowances honestly.", "Time Study", "Allowances are policy — agree them with the team, don't bury them."),
    ("Improve", "Use the standard time for staffing, costing, and line balancing (module in the web app, Phase 2).", "Time Study", "A standard from one operator on one shift is a sample, not a law."),
    ("Control", "For new products/operators, plan capacity with the learning-curve projection, then re-study.", "Learning Curve", "Re-time after the curve flattens; keep the standard current."),
])

# ---------------------------------------------------------------- Time Study
ws = wb.create_sheet("Time Study")
title(ws, "A1", "Time study — observed to standard time")
put_label(ws, "A3", "Allowances (% of normal time)", bold=True, size=9)
put_label(ws, "A4", "Personal"); put_input(ws, "B4", 5)
put_label(ws, "A5", "Fatigue");  put_input(ws, "B5", 4)
put_label(ws, "A6", "Delay");    put_input(ws, "B6", 3)
put_label(ws, "A7", "Total PFD")
put_formula(ws, "B7", '=IF(COUNT($B$4:$B$6)=0,"",SUM($B$4:$B$6))')

HDR = 9; FIRST = 10; LAST = 21  # 12 element rows
header_row(ws, HDR, 1, ["#", "Element description", "Obs1", "Obs2", "Obs3", "Obs4", "Obs5",
                        "Obs6", "Obs7", "Obs8", "Rating %", "Avg observed", "Normal time"],
           widths=[4, 30, 7, 7, 7, 7, 7, 7, 7, 7, 8, 10, 10])
TEX = [
    ("Pick and place housing", [0.42, 0.45, 0.40, 0.44, 0.43], 105),
    ("Insert seal and press",  [0.61, 0.58, 0.63, 0.60, 0.62, 0.59], 100),
    ("Drive 4 screws",         [0.88, 0.92, 0.85, 0.90, 0.91], 95),
    ("Function test",          [0.35, 0.33, 0.36, 0.34], 100),
    ("Label and pack",         [0.28, 0.30, 0.27, 0.29, 0.28], 110),
]
for i in range(12):
    r = FIRST + i
    if i < len(TEX):
        desc, obs, rating = TEX[i]
        put_input(ws, f"A{r}", i + 1)
        put_input(ws, f"B{r}", desc)
        for j in range(8):
            put_input(ws, f"{chr(ord('C')+j)}{r}", obs[j] if j < len(obs) else "")
        put_input(ws, f"K{r}", rating)
    else:
        for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
            put_input(ws, f"{col}{r}", "")
    put_formula(ws, f"L{r}", f'=IF(COUNT(C{r}:J{r})=0,"",AVERAGE(C{r}:J{r}))')
    put_formula(ws, f"M{r}", f'=IF(OR(NOT(ISNUMBER($L{r})),NOT(ISNUMBER($K{r}))),"",$L{r}*$K{r}/100)')

put_label(ws, "L23", "Normal time (sum)", bold=True, size=9)
put_formula(ws, "M23", '=IF(COUNT($M$10:$M$21)=0,"",SUM($M$10:$M$21))')
put_label(ws, "L24", "Standard time", bold=True, size=9)
put_formula(ws, "M24", '=IF(OR($M$23="",$B$7=""),"",$M$23*(1+$B$7/100))')
note(ws, "A26", "Normal = avg observed x rating/100 per element. Standard = sum of normal x (1 + PFD/100). Times in minutes. The web app computes the same numbers from logged observations.")

# ---------------------------------------------------------------- Learning Curve
ws = wb.create_sheet("Learning Curve")
title(ws, "A1", "Learning curve (Wright model)")
put_label(ws, "A3", "Unit-1 time (T1, minutes)", bold=True, size=9)
put_input(ws, "C3", 3.0)
put_label(ws, "A4", "Suggested from Time Study standard:", size=9)
put_link(ws, "C4", "='Time Study'!$M$24")
put_label(ws, "A5", "Learning rate % (e.g. 90 = each doubling takes 90% of prior time)", size=9)
put_input(ws, "C5", 90)
put_label(ws, "A6", "Exponent b = LN(rate)/LN(2)", size=9)
put_formula(ws, "C6", '=IF(OR(NOT(ISNUMBER($C$5)),$C$5<=0),"",LN($C$5/100)/LN(2))')

HDR = 8; FIRST = 9; LAST = 40  # units 1..32
header_row(ws, HDR, 1, ["Unit n", "Time per unit (min)", "Cumulative total (min)"],
           widths=[8, 16, 18])
for i in range(32):
    r = FIRST + i
    put_label(ws, f"A{r}", i + 1)
    put_formula(ws, f"B{r}", f'=IF(OR(NOT(ISNUMBER($C$3)),$C$6=""),"",$C$3*$A{r}^$C$6)')
    put_formula(ws, f"C{r}", f'=IF(ISNUMBER($B{r}),SUM($B$9:$B{r}),"")')

ch = LineChart(); ch.title = "Time per unit"; ch.height = 9; ch.width = 20
s = Series(Reference(ws, min_col=2, min_row=HDR, max_row=LAST), title_from_data=True)
s.smooth = False
ch.series.append(s)
ch.set_categories(Reference(ws, min_col=1, min_row=FIRST, max_row=LAST))
ws.add_chart(ch, "E3")
note(ws, "A42", "T_n = T1 x n^b with b = LN(rate)/LN(2). Type your own T1 in the yellow cell (the green cell shows the Time Study suggestion). Use for ramp-up planning; re-study once the curve flattens.")

wb.save("../IE_Toolkit.xlsx")
print("IE_Toolkit.xlsx written")
