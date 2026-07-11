"""Build PE_Toolkit.xlsx — Process Engineering companion workbook."""
import random
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, Series
from openpyxl.worksheet.datavalidation import DataValidation
from common import (build_guide, build_setup, build_workflow, header_row, put_input,
                    put_formula, put_link, put_label, note, title, SETUP)

random.seed(42)

wb = Workbook()
wb.remove(wb.active)

build_guide(wb, "PE Toolkit (Process Engineering)", [
    ("Setup", "Industry & process type — drives terminology and thresholds on every sheet."),
    ("Workflow - Start Here", "Guided DMAIC sequence, mirroring the web app's playbook."),
    ("SPC X-bar R", "Subgrouped control chart: limits from A2/D3/D4 constants, WE1 & WE4 flags, X-bar and R charts."),
    ("I-MR", "Individuals & moving range chart for one-at-a-time measurements."),
    ("Capability", "Cp/Cpk (within, R-bar/d2) and Pp/Ppk (overall) pulled live from the SPC sheet."),
    ("OEE Log", "Daily availability x performance x quality with below-target flags and trend chart."),
], honesty_notes=[
    "This workbook checks Western Electric rules 1 (beyond 3-sigma) and 4 (8 consecutive one side). The web app checks all four WE rules on every entry.",
    "Charts reference the full 30-row input range; rows you have not filled yet can render as gaps or zeros depending on your spreadsheet app - a charting artifact, not data.",
    "Control limits assume rational subgroups and approximate normality; with fewer than ~20 subgroups treat limits as preliminary.",
])
build_setup(wb)
build_workflow(wb, [
    ("Define", "Write the problem statement: one process, one measurable characteristic, one baseline metric.", "—", "Vague problem = unmeasurable improvement."),
    ("Measure", "Log at least 20 subgroups on SPC X-bar R (or 20+ individuals on I-MR).", "SPC X-bar R / I-MR", "Log as it happens, not from memory. Keep subgroup size constant."),
    ("Analyze", "Check WE flags and the Capability sheet against your industry Cpk threshold.", "Capability", "Cpk below threshold means the process can't hold spec even in control — fix the process, not the operator."),
    ("Improve", "Implement the change; keep logging the SAME characteristic the SAME way.", "SPC X-bar R", "Changing the measurement method mid-study invalidates the comparison."),
    ("Control", "Recalculate limits after the change proves stable; keep the chart alive.", "SPC X-bar R", "A control chart nobody updates is decoration."),
])

# ---------------------------------------------------------------- SPC X-bar R
ws = wb.create_sheet("SPC X-bar R")
title(ws, "A1", "X-bar & R control chart (subgrouped)")
put_label(ws, "A3", "Subgroup size n (2-5 measurement columns used)", bold=True, size=9)
put_input(ws, "B3", 5)
dvn = DataValidation(type="list", formula1='"2,3,4,5"', allow_blank=False)
ws.add_data_validation(dvn); dvn.add(ws["B3"])
note(ws, "A4", "Type measurements in yellow x1..x5 (leave extra columns empty if n<5). Everything else computes.")

HDR = 8; FIRST = 9; LAST = 38  # 30 data rows
header_row(ws, HDR, 1, ["Subgroup", "Date", "x1", "x2", "x3", "x4", "x5",
                        "Mean", "Range", "CL", "UCLx", "LCLx", "UCLr", "LCLr", "WE1", "WE4"],
           widths=[9, 11, 7, 7, 7, 7, 7, 8, 8, 8, 8, 8, 8, 8, 6, 6])

# example data: 20 subgroups ~N(10, 0.01), subgroup 15 shifted +0.05 to trip WE1
for i in range(20):
    r = FIRST + i
    put_input(ws, f"A{r}", i + 1)
    put_input(ws, f"B{r}", f"2026-06-{i+1:02d}")
    shift = 0.05 if i == 14 else 0.0
    for j in range(5):
        put_input(ws, f"{chr(ord('C')+j)}{r}", round(10.0 + shift + random.gauss(0, 0.01), 3))
for i in range(20, 30):  # empty input rows ready for the user
    r = FIRST + i
    put_input(ws, f"A{r}", "")
    put_input(ws, f"B{r}", "")
    for j in range(5):
        put_input(ws, f"{chr(ord('C')+j)}{r}", "")

for r in range(FIRST, LAST + 1):
    put_formula(ws, f"H{r}", f'=IF(COUNT(C{r}:G{r})=0,"",AVERAGE(C{r}:G{r}))')
    put_formula(ws, f"I{r}", f'=IF(COUNT(C{r}:G{r})=0,"",MAX(C{r}:G{r})-MIN(C{r}:G{r}))')
    put_formula(ws, f"J{r}", f'=IF(ISNUMBER($H{r}),$S$3,"")')
    put_formula(ws, f"K{r}", f'=IF(ISNUMBER($H{r}),$S$9,"")')
    put_formula(ws, f"L{r}", f'=IF(ISNUMBER($H{r}),$S$10,"")')
    put_formula(ws, f"M{r}", f'=IF(ISNUMBER($I{r}),$S$11,"")')
    put_formula(ws, f"N{r}", f'=IF(ISNUMBER($I{r}),$S$12,"")')
    put_formula(ws, f"O{r}", f'=IF(ISNUMBER($H{r}),IF(OR($H{r}>$S$9,$H{r}<$S$10),"WE1",""),"")')
    if r >= FIRST + 7:
        lo = r - 7
        put_formula(ws, f"P{r}",
            f'=IF(COUNT(H{lo}:H{r})=8,IF(OR(COUNTIF(H{lo}:H{r},">"&$S$3)=8,COUNTIF(H{lo}:H{r},"<"&$S$3)=8),"WE4",""),"")')

# stats block
put_label(ws, "R2", "Statistics", bold=True, size=9)
stats = [("X-double-bar", '=IF(COUNT($H$9:$H$38)=0,"",AVERAGE($H$9:$H$38))'),
         ("R-bar", '=IF(COUNT($I$9:$I$38)=0,"",AVERAGE($I$9:$I$38))'),
         ("A2", '=INDEX($S$17:$S$25,MATCH($B$3,$R$17:$R$25,0))'),
         ("D3", '=INDEX($T$17:$T$25,MATCH($B$3,$R$17:$R$25,0))'),
         ("D4", '=INDEX($U$17:$U$25,MATCH($B$3,$R$17:$R$25,0))'),
         ("d2", '=INDEX($V$17:$V$25,MATCH($B$3,$R$17:$R$25,0))'),
         ("UCLx", '=IF(OR($S$3="",$S$4=""),"",$S$3+$S$5*$S$4)'),
         ("LCLx", '=IF(OR($S$3="",$S$4=""),"",$S$3-$S$5*$S$4)'),
         ("UCLr", '=IF($S$4="","",$S$7*$S$4)'),
         ("LCLr", '=IF($S$4="","",$S$6*$S$4)'),
         ("sigma within (R-bar/d2)", '=IF($S$4="","",$S$4/$S$8)')]
for i, (lab, f) in enumerate(stats):
    put_label(ws, f"R{3+i}", lab, size=9)
    put_formula(ws, f"S{3+i}", f)
ws.column_dimensions["R"].width = 20

put_label(ws, "R16", "Control chart constants (Montgomery, Introduction to SQC)", bold=True, size=8)
header_row(ws, 16, 18, ["n", "A2", "D3", "D4", "d2"], widths=[5, 7, 7, 7, 7])
CONSTS = [(2,1.88,0,3.267,1.128),(3,1.023,0,2.574,1.693),(4,0.729,0,2.282,2.059),
          (5,0.577,0,2.114,2.326),(6,0.483,0,2.004,2.534),(7,0.419,0.076,1.924,2.704),
          (8,0.373,0.136,1.864,2.847),(9,0.337,0.184,1.816,2.970),(10,0.308,0.223,1.777,3.078)]
for i, row in enumerate(CONSTS):
    for j, v in enumerate(row):
        c = ws.cell(row=17 + i, column=18 + j, value=v)
        c.font = c.font.copy(name="Arial", size=8)

# X-bar chart
ch = LineChart(); ch.title = "X-bar chart"; ch.height = 8; ch.width = 22; ch.style = 2
cats = Reference(ws, min_col=1, min_row=FIRST, max_row=LAST)
for col, name in [(8, "Mean"), (10, "CL"), (11, "UCL"), (12, "LCL")]:
    s = Series(Reference(ws, min_col=col, min_row=HDR, max_row=LAST), title_from_data=True)
    s.smooth = False
    ch.series.append(s)
ch.set_categories(cats)
ws.add_chart(ch, "A41")
note(ws, "A40", "Charts plot all 30 input rows. Unfilled rows can show as gaps or zeros depending on your app — charting artifact, not data. WE1 = point beyond 3-sigma; WE4 = 8 consecutive one side of CL (the web app also checks WE2/WE3).")

ch2 = LineChart(); ch2.title = "R chart"; ch2.height = 8; ch2.width = 22; ch2.style = 2
for col, name in [(9, "Range"), (13, "UCLr"), (14, "LCLr")]:
    s = Series(Reference(ws, min_col=col, min_row=HDR, max_row=LAST), title_from_data=True)
    s.smooth = False
    ch2.series.append(s)
ch2.set_categories(cats)
ws.add_chart(ch2, "A58")

# ---------------------------------------------------------------- I-MR
ws = wb.create_sheet("I-MR")
title(ws, "A1", "Individuals & moving range chart")
note(ws, "A3", "For one-at-a-time measurements (no rational subgroups). Type values in yellow; MR, limits, and flags compute.")
HDR = 5; FIRST = 6; LAST = 35
header_row(ws, HDR, 1, ["Date", "Value", "MR", "CL", "UCL", "LCL", "WE1"],
           widths=[11, 9, 8, 8, 8, 8, 6])
vals = [round(10.0 + random.gauss(0, 0.02), 3) for _ in range(19)] + [10.15]  # last point spikes
for i in range(30):
    r = FIRST + i
    if i < 20:
        put_input(ws, f"A{r}", f"2026-06-{i+1:02d}")
        put_input(ws, f"B{r}", vals[i])
    else:
        put_input(ws, f"A{r}", ""); put_input(ws, f"B{r}", "")
for r in range(FIRST, LAST + 1):
    if r > FIRST:
        put_formula(ws, f"C{r}", f'=IF(AND(ISNUMBER($B{r}),ISNUMBER($B{r-1})),ABS($B{r}-$B{r-1}),"")')
    put_formula(ws, f"D{r}", f'=IF(ISNUMBER($B{r}),$J$4,"")')
    put_formula(ws, f"E{r}", f'=IF(ISNUMBER($B{r}),$J$7,"")')
    put_formula(ws, f"F{r}", f'=IF(ISNUMBER($B{r}),$J$8,"")')
    put_formula(ws, f"G{r}", f'=IF(ISNUMBER($B{r}),IF(OR($B{r}>$J$7,$B{r}<$J$8),"WE1",""),"")')
put_label(ws, "I3", "Statistics", bold=True, size=9)
imr = [("X-bar", '=IF(COUNT($B$6:$B$35)=0,"",AVERAGE($B$6:$B$35))'),
       ("MR-bar", '=IF(COUNT($C$7:$C$35)=0,"",AVERAGE($C$7:$C$35))'),
       ("sigma (MR-bar/1.128)", '=IF($J$5="","",$J$5/1.128)'),
       ("UCL (X+3s)", '=IF(OR($J$4="",$J$6=""),"",$J$4+3*$J$6)'),
       ("LCL (X-3s)", '=IF(OR($J$4="",$J$6=""),"",$J$4-3*$J$6)'),
       ("UCL-MR (3.267*MR-bar)", '=IF($J$5="","",3.267*$J$5)')]
for i, (lab, f) in enumerate(imr):
    put_label(ws, f"I{4+i}", lab, size=9)
    put_formula(ws, f"J{4+i}", f)
ws.column_dimensions["I"].width = 22
ch = LineChart(); ch.title = "Individuals chart"; ch.height = 8; ch.width = 22
for col in [2, 4, 5, 6]:
    s = Series(Reference(ws, min_col=col, min_row=HDR, max_row=LAST), title_from_data=True)
    s.smooth = False
    ch.series.append(s)
ch.set_categories(Reference(ws, min_col=1, min_row=FIRST, max_row=LAST))
ws.add_chart(ch, "A38")
note(ws, "A37", "Unfilled rows may render as gaps/zeros — charting artifact, not data.")

# ---------------------------------------------------------------- Capability
ws = wb.create_sheet("Capability")
title(ws, "A1", "Process capability — pulled live from 'SPC X-bar R'")
put_label(ws, "A3", "Spec limits (inputs)", bold=True, size=9)
put_label(ws, "A4", "LSL"); put_input(ws, "B4", 9.94)
put_label(ws, "A5", "USL"); put_input(ws, "B5", 10.06)
put_label(ws, "A7", "From SPC sheet (green = cross-sheet links)", bold=True, size=9)
put_label(ws, "A8", "Process mean (X-double-bar)")
put_link(ws, "B8", "='SPC X-bar R'!$S$3")
put_label(ws, "A9", "sigma within (R-bar/d2)")
put_link(ws, "B9", "='SPC X-bar R'!$S$13")
put_label(ws, "A10", "sigma overall (sample s of all obs)")
put_link(ws, "B10", "=IF(COUNT('SPC X-bar R'!$C$9:$G$38)<2,\"\",STDEV('SPC X-bar R'!$C$9:$G$38))")
put_label(ws, "A11", "Cpk threshold (from Setup)")
put_link(ws, "B11", f"={SETUP['cpk']}")

put_label(ws, "A13", "Indices", bold=True, size=9)
caps = [("Cp  (within)",  '=IF(OR($B$4="",$B$5="",$B$9="",$B$9=0),"",($B$5-$B$4)/(6*$B$9))'),
        ("Cpk (within)",  '=IF(OR($B$4="",$B$5="",$B$8="",$B$9="",$B$9=0),"",MIN(($B$5-$B$8)/(3*$B$9),($B$8-$B$4)/(3*$B$9)))'),
        ("Pp  (overall)", '=IF(OR($B$4="",$B$5="",$B$10="",$B$10=0),"",($B$5-$B$4)/(6*$B$10))'),
        ("Ppk (overall)", '=IF(OR($B$4="",$B$5="",$B$8="",$B$10="",$B$10=0),"",MIN(($B$5-$B$8)/(3*$B$10),($B$8-$B$4)/(3*$B$10)))')]
for i, (lab, f) in enumerate(caps):
    put_label(ws, f"A{14+i}", lab)
    c = put_formula(ws, f"B{14+i}", f); c.number_format = "0.00"
put_label(ws, "A19", "Verdict")
put_formula(ws, "B19", '=IF($B$15="","log SPC data first",IF($B$15>=$B$11,"CAPABLE vs threshold","NOT CAPABLE - below threshold"))')
note(ws, "A21", "Cp/Cpk use within-subgroup sigma (R-bar/d2): short-term, what the process CAN do. Pp/Ppk use overall s: long-term, what it DID do. Indices assume approximate normality. In the web app a Cpk drop below threshold raises an alert automatically.")
ws.merge_cells("A21:F22")
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 16

# ---------------------------------------------------------------- OEE Log
ws = wb.create_sheet("OEE Log")
title(ws, "A1", "OEE daily log")
put_label(ws, "A3", "OEE target (from Setup)")
c = put_link(ws, "B3", f"={SETUP['oee']}"); c.number_format = "0%"
HDR = 6; FIRST = 7; LAST = 36
header_row(ws, HDR, 1, ["Date", "Planned min", "Run min", "Ideal cycle (min/unit)", "Total count",
                        "Good count", "Availability", "Performance", "Quality", "OEE", "Target", "Below?"],
           widths=[11, 11, 9, 13, 10, 10, 11, 11, 9, 9, 8, 8])
ex = []
for i in range(14):
    dip = i in (9, 10, 11)  # 3 consecutive bad days -> alert pattern
    runtime = 380 if dip else 450
    total = 330 if dip else 430
    good = total - (18 if dip else 6)
    ex.append((f"2026-06-{i+1:02d}", 480, runtime, 1.0, total, good))
for i in range(30):
    r = FIRST + i
    if i < len(ex):
        d, pl, rt, ic, tot, gd = ex[i]
        put_input(ws, f"A{r}", d); put_input(ws, f"B{r}", pl); put_input(ws, f"C{r}", rt)
        put_input(ws, f"D{r}", ic); put_input(ws, f"E{r}", tot); put_input(ws, f"F{r}", gd)
    else:
        for col in "ABCDEF":
            put_input(ws, f"{col}{r}", "")
for r in range(FIRST, LAST + 1):
    a = put_formula(ws, f"G{r}", f'=IF(OR(NOT(ISNUMBER($B{r})),NOT(ISNUMBER($C{r})),$B{r}=0),"",$C{r}/$B{r})')
    p = put_formula(ws, f"H{r}", f'=IF(OR(NOT(ISNUMBER($C{r})),NOT(ISNUMBER($D{r})),NOT(ISNUMBER($E{r})),$C{r}=0),"",$D{r}*$E{r}/$C{r})')
    q = put_formula(ws, f"I{r}", f'=IF(OR(NOT(ISNUMBER($E{r})),NOT(ISNUMBER($F{r})),$E{r}=0),"",$F{r}/$E{r})')
    o = put_formula(ws, f"J{r}", f'=IF(OR($G{r}="",$H{r}="",$I{r}=""),"",$G{r}*$H{r}*$I{r})')
    t = put_formula(ws, f"K{r}", f'=IF(ISNUMBER($J{r}),$B$3,"")')
    put_formula(ws, f"L{r}", f'=IF(ISNUMBER($J{r}),IF($J{r}<$B$3,"BELOW",""),"")')
    for cc in (a, p, q, o, t):
        cc.number_format = "0.0%"
ch = LineChart(); ch.title = "OEE trend vs target"; ch.height = 8; ch.width = 22
for col in [10, 11]:
    s = Series(Reference(ws, min_col=col, min_row=HDR, max_row=LAST), title_from_data=True)
    s.smooth = False
    ch.series.append(s)
ch.set_categories(Reference(ws, min_col=1, min_row=FIRST, max_row=LAST))
ws.add_chart(ch, "A39")
note(ws, "A38", "3+ consecutive BELOW rows = the pattern the web app alerts on automatically. Unfilled rows may render as gaps/zeros — charting artifact.")

wb.save("../PE_Toolkit.xlsx")
print("PE_Toolkit.xlsx written")
