"""Shared helpers for the Procyra companion workbooks.

Conventions (per product spec):
  - INPUT cells: yellow fill + blue bold text
  - FORMULA cells: black text, never hardcoded results
  - CROSS-SHEET links: green text
  - Guide sheet with legend + reuse instructions on every workbook
  - Setup sheet: industry/process-type selection drives terminology via INDEX/MATCH
  - Compatibility: no XLOOKUP/FILTER/SORT/UNIQUE/SEQUENCE anywhere
"""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

ARIAL = "Arial"
BLUE = "0000FF"; GREEN = "008000"; BLACK = "000000"; WHITE = "FFFFFF"
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
HEADER_FILL = PatternFill("solid", fgColor="22262B")
SECTION_FILL = PatternFill("solid", fgColor="DFE2DC")
THIN = Border(*(Side(style="thin", color="B0B4AD"),) * 4)

def title(ws, cell, text):
    c = ws[cell]; c.value = text
    c.font = Font(name=ARIAL, bold=True, size=14, color=BLACK)

def header_row(ws, row, col_start, labels, widths=None):
    for i, lab in enumerate(labels):
        c = ws.cell(row=row, column=col_start + i, value=lab)
        c.font = Font(name=ARIAL, bold=True, size=9, color=WHITE)
        c.fill = HEADER_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = THIN
        if widths:
            ws.column_dimensions[c.column_letter].width = widths[i]

def put_input(ws, cell, value):
    c = ws[cell]; c.value = value
    c.fill = YELLOW_FILL
    c.font = Font(name=ARIAL, bold=True, color=BLUE)
    c.border = THIN
    return c

def put_formula(ws, cell, formula):
    c = ws[cell]; c.value = formula
    c.font = Font(name=ARIAL, color=BLACK)
    c.border = THIN
    return c

def put_link(ws, cell, formula):
    """Cross-sheet link — green text per convention."""
    c = ws[cell]; c.value = formula
    c.font = Font(name=ARIAL, color=GREEN)
    c.border = THIN
    return c

def put_label(ws, cell, text, bold=False, size=10, wrap=False, italic=False, color=BLACK):
    c = ws[cell]; c.value = text
    c.font = Font(name=ARIAL, bold=bold, size=size, italic=italic, color=color)
    if wrap:
        c.alignment = Alignment(wrap_text=True, vertical="top")
    return c

def note(ws, cell, text):
    return put_label(ws, cell, text, size=8, italic=True, color="5A6472")

INDUSTRIES = ["General manufacturing", "Automotive", "Aerospace & defense",
              "Food & beverage", "Pharma / medical device", "Electronics",
              "Warehouse & logistics", "Construction"]
PROCESS_TYPES = ["Discrete assembly", "Batch production", "Continuous process",
                 "Job shop", "Service / transactional"]
# columns: Defect term | Line term | Unit term | Quality standard | Cpk threshold | OEE target | NC repeat threshold | RPN action
TERM_TABLE = [
    ["General manufacturing", "Defect", "Line", "Unit", "ISO 9001", 1.33, 0.85, 3, 100],
    ["Automotive", "Defect", "Line", "Unit", "IATF 16949", 1.33, 0.85, 3, 100],
    ["Aerospace & defense", "Non-conformity", "Line", "Unit", "AS9100", 1.67, 0.85, 2, 80],
    ["Food & beverage", "Deviation", "Processing line", "Batch", "ISO 9001 / FSSC 22000", 1.33, 0.85, 3, 100],
    ["Pharma / medical device", "Deviation", "Line", "Lot", "ISO 13485 / 21 CFR 820", 1.67, 0.85, 2, 80],
    ["Electronics", "Defect", "SMT line", "Board", "ISO 9001", 1.33, 0.85, 3, 100],
    ["Warehouse & logistics", "Error", "Pick line", "Order", "ISO 9001", 1.33, 0.75, 3, 100],
    ["Construction", "Punch item", "Crew", "Task", "ISO 9001", 1.33, 0.75, 3, 100],
]

def build_setup(wb):
    ws = wb.create_sheet("Setup")
    title(ws, "A1", "Setup — industry & process type")
    put_label(ws, "A2", "Pick your industry and process type; terminology and defaults below adapt automatically and feed every other sheet.", size=9, wrap=True)
    ws.merge_cells("A2:F2")
    put_label(ws, "A4", "Industry", bold=True)
    put_input(ws, "B4", "Automotive")
    put_label(ws, "A5", "Process type", bold=True)
    put_input(ws, "B5", "Discrete assembly")
    dv1 = DataValidation(type="list", formula1='"' + ",".join(INDUSTRIES) + '"', allow_blank=False)
    dv2 = DataValidation(type="list", formula1='"' + ",".join(PROCESS_TYPES) + '"', allow_blank=False)
    ws.add_data_validation(dv1); dv1.add(ws["B4"])
    ws.add_data_validation(dv2); dv2.add(ws["B5"])

    put_label(ws, "A7", "Adapted terminology & defaults (formulas — do not type here)", bold=True, size=9)
    labels = ["Defect term", "Line term", "Unit term", "Quality standard",
              "Cpk threshold", "OEE target", "NC repeat threshold", "RPN action threshold"]
    for i, lab in enumerate(labels):
        r = 8 + i
        put_label(ws, f"A{r}", lab)
        put_formula(ws, f"B{r}", f"=INDEX($E$21:$L$28,MATCH($B$4,$D$21:$D$28,0),{i+1})")

    put_label(ws, "D19", "Terminology dictionary (edit to add industries)", bold=True, size=9)
    header_row(ws, 20, 4, ["Industry", "Defect term", "Line term", "Unit term", "Quality standard",
                           "Cpk thr", "OEE target", "NC repeat", "RPN action"],
               widths=[24, 14, 14, 10, 22, 8, 10, 9, 10])
    for ri, row in enumerate(TERM_TABLE):
        for ci, v in enumerate(row):
            c = ws.cell(row=21 + ri, column=4 + ci, value=v)
            c.font = Font(name=ARIAL, size=9, bold=True, color=BLUE)
            c.fill = YELLOW_FILL
            c.border = THIN
    note(ws, "A17", "Other sheets read B8:B15 (green cells there = links to here).")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 24
    return ws

# Cell map for cross-sheet links into Setup
SETUP = {
    "defect": "Setup!$B$8", "line": "Setup!$B$9", "unit": "Setup!$B$10",
    "standard": "Setup!$B$11", "cpk": "Setup!$B$12", "oee": "Setup!$B$13",
    "nc_repeat": "Setup!$B$14", "rpn": "Setup!$B$15",
}

def build_guide(wb, workbook_name, sheet_notes, honesty_notes=None):
    ws = wb.create_sheet("Guide", 0)
    title(ws, "A1", f"Procyra — {workbook_name}")
    put_label(ws, "A2", "Companion workbook to the Procyra web app. Same logic, same Phase 1 modules — usable offline and shareable anywhere.", size=9, wrap=True)
    ws.merge_cells("A2:E2")

    put_label(ws, "A4", "Color legend", bold=True)
    put_input(ws, "A5", "Input cell — type here (yellow fill, blue bold)")
    ws.merge_cells("A5:D5")
    put_formula(ws, "A6", "Formula cell — never overwrite (black text)")
    ws.merge_cells("A6:D6")
    put_link(ws, "A7", "Cross-sheet link (green text)")
    ws.merge_cells("A7:D7")

    put_label(ws, "A9", "How to reuse this workbook", bold=True)
    steps = [
        "1. Go to Setup and pick your industry & process type — labels and defaults adapt everywhere.",
        "2. Every sheet ships with a documented example dataset in the input cells. Study it, then replace it with your data.",
        "3. Only ever type in yellow cells. Black cells are formulas; green cells pull from another sheet.",
        "4. To extend a table, insert rows INSIDE the existing range and copy the formula row down.",
        "5. Compatibility: only classic functions (INDEX/MATCH, SUMIFS, SUMPRODUCT, RANK) — works in old Excel and LibreOffice.",
    ]
    for i, s_ in enumerate(steps):
        put_label(ws, f"A{10+i}", s_, size=9)

    r = 16
    put_label(ws, f"A{r}", "Sheets in this workbook", bold=True)
    for i, (name, desc) in enumerate(sheet_notes):
        put_label(ws, f"A{r+1+i}", name, bold=True, size=9)
        put_label(ws, f"B{r+1+i}", desc, size=9)
    r = r + 2 + len(sheet_notes)
    if honesty_notes:
        put_label(ws, f"A{r}", "What this workbook does NOT do (honest limits)", bold=True)
        for i, n_ in enumerate(honesty_notes):
            put_label(ws, f"A{r+1+i}", "• " + n_, size=9)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 90
    return ws

def build_workflow(wb, rows):
    """'Start Here' sheet mirroring the app's guided DMAIC playbook."""
    ws = wb.create_sheet("Workflow - Start Here")
    title(ws, "A1", "Guided workflow — Reduce defects (DMAIC)")
    put_label(ws, "A2", "Mirrors the Procyra web app's guided playbook. Work top to bottom. The app suggests — your team decides; record decisions in the last column.", size=9, wrap=True)
    ws.merge_cells("A2:E2")
    header_row(ws, 4, 1, ["Phase", "What to do", "Sheet to use", "Watch out for", "Team decision / notes (type here)"],
               widths=[10, 46, 20, 40, 40])
    for i, (phase, what, sheet, watch) in enumerate(rows):
        r = 5 + i
        put_label(ws, f"A{r}", phase, bold=True, size=9)
        put_label(ws, f"B{r}", what, size=9, wrap=True)
        put_label(ws, f"C{r}", sheet, size=9)
        put_label(ws, f"D{r}", watch, size=9, wrap=True)
        put_input(ws, f"E{r}", "")
        ws.row_dimensions[r].height = 42
    return ws
